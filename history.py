import subprocess
import openpyxl

def git_log_to_excel(output_file):
    # Run git log -p command and capture the output
    git_log_output = subprocess.check_output(['','git', 'log', '-p'], text=True)

    # Split the log into individual commits
    commits = git_log_output.split('\ncommit ')

    # Create a new Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active

    # Write headers to the Excel file
    headers = ['Author', 'Date', 'Changes']
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)

    # Iterate through each commit and write data to the Excel file
    for commit in commits[1:]:
        commit_lines = commit.split('\n')
        commit_hash = commit_lines[0].strip()

        # Extract author and date information
        author_line = next(line for line in commit_lines if line.startswith('Author:'))
        author = author_line.split(':')[1].strip()

        date_line = next(line for line in commit_lines if line.startswith('Date:'))
        date = date_line.split(':')[1].strip()

        # Extract changes
        changes = '\n'.join(commit_lines[8:])  # Assuming that the first 6 lines are metadata

        # Write data to the Excel file
        row_num = ws.max_row + 1
        #ws.cell(row=row_num, column=1, value=commit_hash)
        ws.cell(row=row_num, column=1, value=author)
        ws.cell(row=row_num, column=2, value=date)
        ws.cell(row=row_num, column=3, value=changes)

    # Save the Excel file
    wb.save(output_file)

if __name__ == "__main__":
    git_log_to_excel("git_log_output.xlsx")

